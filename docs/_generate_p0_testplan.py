"""Generate AISI Manual Test Plan P0 Excel. Run: python docs/_generate_p0_testplan.py"""
from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "AISI_Manual_Test_Plan_P0.xlsx"

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="1B4F72")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
title_font = Font(bold=True, name="Calibri", size=16, color="1B4F72")
subtitle_font = Font(bold=True, name="Calibri", size=12, color="2874A6")
section_fill = PatternFill("solid", fgColor="D4E6F1")
p0_fill = PatternFill("solid", fgColor="F5B7B1")
happy_fill = PatternFill("solid", fgColor="D5F5E3")
neg_fill = PatternFill("solid", fgColor="FADBD8")
smoke_fill = PatternFill("solid", fgColor="FCF3CF")
pass_fill = PatternFill("solid", fgColor="82E0AA")
fail_fill = PatternFill("solid", fgColor="F1948A")
blocked_fill = PatternFill("solid", fgColor="F7DC6F")
yes_fill = PatternFill("solid", fgColor="D5F5E3")
no_fill = PatternFill("solid", fgColor="F5B7B1")
scope_fill = PatternFill("solid", fgColor="FCF3CF")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin


def style_data_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)


def set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_cases() -> list[tuple]:
    cases: list[tuple] = []

    def add(cid, modul, judul, tipe, role, pre, steps, expected, data):
        cases.append(
            (cid, modul, judul, tipe, "P0", role, pre, steps, expected, data, "Not Run", "", "", "")
        )

    # AUTH
    add(
        "AUTH-01",
        "AUTH",
        "Login sukses SUPERADMIN",
        "Happy",
        "SUPERADMIN",
        "API+FE jalan; DB seed; logout / jendela Incognito",
        "1. Buka http://localhost:5173/login\n"
        "2. Isi email & password Superadmin\n"
        "3. Klik Login / Masuk\n"
        "4. Amati URL dan konten dashboard",
        "Redirect ke /dashboard; sidebar Superadmin tampil; tidak ada error toast; refresh tetap login",
        "fuadinaqi@gmail.com / !Superadmin123",
    )
    add(
        "AUTH-02",
        "AUTH",
        "Login gagal password salah",
        "Negatif",
        "Semua",
        "Di halaman /login",
        "1. Isi email valid (mis. Superadmin)\n2. Isi password salah\n3. Klik Login",
        "Tetap di /login; pesan error credentials tampil; tidak masuk dashboard",
        "email valid + password salah",
    )
    add(
        "AUTH-03",
        "AUTH",
        "Login semua role seed (smoke)",
        "Smoke",
        "Semua role",
        "5 akun seed tersedia",
        "Untuk tiap role (SA, Admin, PJ, Pembina, Anggota):\n"
        "1. Logout jika perlu\n2. Login dengan kredensial role\n3. Verifikasi landing /dashboard",
        "Semua role berhasil login ke /dashboard tanpa error",
        "Lihat sheet 02_Akun_Seed",
    )
    add(
        "AUTH-04",
        "AUTH",
        "Logout dan proteksi route privat",
        "Happy",
        "SUPERADMIN",
        "Sudah login",
        "1. Klik Logout\n2. Pastikan diarahkan ke /login\n"
        "3. Ketik URL /dashboard di address bar\n4. Tekan Enter",
        "Setelah logout di /login; akses /dashboard tanpa token redirect ke /login",
        "—",
    )
    add(
        "AUTH-05",
        "AUTH",
        "Set password dari undangan (token seed)",
        "Happy",
        "Public",
        "Token seed invite masih valid di DB, atau buat undangan baru lalu ambil token dari log API",
        "1. Buka /set-password?token=<SEED_INVITE_TOKEN atau token baru>\n"
        "2. Isi password baru sesuai aturan\n3. Konfirmasi password\n4. Submit\n"
        "5. Login dengan email undangan + password baru",
        "Password tersimpan; user bisa login; role sesuai undangan ter-assign",
        "Token: 00000000-0000-4000-8000-000000000001 (atau dari log)",
    )
    add(
        "AUTH-06",
        "AUTH",
        "Set password token invalid/expired",
        "Negatif",
        "Public",
        "—",
        "1. Buka /set-password?token=invalid-token-xyz\n2. Coba isi & submit password",
        "Pesan error token tidak valid/kadaluarsa; password tidak tersimpan",
        "token=invalid-token-xyz",
    )
    add(
        "AUTH-07",
        "AUTH",
        "Accept role untuk existing user",
        "Happy",
        "Public / existing user",
        "Token accept-role valid (seed atau undangan ke email existing)",
        "1. Buka /accept-role?token=<token>\n2. Login jika diminta (email harus cocok)\n"
        "3. Klik Terima / Accept peran\n4. Cek RoleSwitcher / profil roles",
        "Role baru ditambahkan; user bisa switch ke role tersebut",
        "SEED_ACCEPT_ROLE_TOKEN atau undangan ke email existing",
    )
    add(
        "AUTH-08",
        "AUTH",
        "Accept role ditolak jika email tidak cocok",
        "Negatif",
        "Public",
        "Token undangan untuk email A; login sebagai email B",
        "1. Login sebagai user B\n"
        "2. Buka /accept-role?token= untuk undangan email A\n3. Coba accept",
        "Ditolak (403/pesan error); role tidak ditambahkan ke user B",
        "Email mismatch",
    )
    add(
        "AUTH-09",
        "AUTH",
        "Forgot & Reset password (smoke)",
        "Smoke",
        "Public",
        "Email provider/log API tersedia",
        "1. Buka /forgot-password\n2. Submit email akun seed\n"
        "3. Ambil link reset dari email/log API\n4. Buka /reset-password?token=...\n"
        "5. Set password baru lalu login",
        "Flow selesai tanpa crash; login password baru berhasil "
        "(atau Skip/Blocked jika email env tidak dikonfigurasi)",
        "Akun seed non-critical",
    )
    add(
        "AUTH-10",
        "AUTH",
        "Redirect aman setelah login (?redirect=)",
        "Happy",
        "SUPERADMIN",
        "Logout",
        "1. Buka /login?redirect=/schools\n2. Login Superadmin\n"
        "3. Ulangi dengan /login?redirect=//evil.com",
        "redirect=/schools → mendarat di /schools; "
        "redirect=//evil.com → fallback /dashboard (anti open-redirect)",
        "redirect aman vs berbahaya",
    )

    # ROLE
    add(
        "ROLE-01",
        "ROLE",
        "Nav loop SUPERADMIN",
        "Smoke",
        "SUPERADMIN",
        "Login Superadmin",
        "1. Login\n2. Klik berurutan semua item sidebar SUPERADMIN\n"
        "3. Di tiap halaman pastikan main content terlihat (bukan blank)\n"
        "4. Catat pageerror di console jika ada",
        "Semua menu terbuka; konten utama terlihat; tidak blank; tidak uncaught exception kritis",
        "Nav: Beranda…Profil (lihat matriks)",
    )
    add(
        "ROLE-02",
        "ROLE",
        "Nav loop ADMIN",
        "Smoke",
        "ADMIN",
        "Login Admin",
        "1. Login Admin\n2. Loop semua item sidebar ADMIN\n"
        "3. Pastikan menu Pengguna TIDAK ada\n4. Konten tiap halaman terlihat",
        "Nav sesuai ADMIN; /users tidak di sidebar; halaman tidak blank",
        "fuadiproject@gmail.com",
    )
    add(
        "ROLE-03",
        "ROLE",
        "Nav loop PJ_SEKOLAH",
        "Smoke",
        "PJ_SEKOLAH",
        "Login PJ",
        "1. Login PJ\n2. Loop menu PJ termasuk /pembina dan /config/ic\n"
        "3. Pastikan tidak ada menu Pengguna",
        "Nav sesuai PJ; konten terlihat",
        "usamah_sman1@gmail.com",
    )
    add(
        "ROLE-04",
        "ROLE",
        "Nav loop PEMBINA",
        "Smoke",
        "PEMBINA",
        "Login Pembina",
        "1. Login Pembina\n2. Pastikan menu Evaluasi ada\n"
        "3. Loop semua item nav PEMBINA\n4. Pastikan Mutabaah tidak di sidebar",
        "Nav sesuai PEMBINA; Evaluasi tersedia; Mutabaah tidak muncul",
        "budi.santoso.sman1.pembina@gmail.com",
    )
    add(
        "ROLE-05",
        "ROLE",
        "Nav loop ANGGOTA",
        "Smoke",
        "ANGGOTA",
        "Login Anggota",
        "1. Login Anggota\n2. Pastikan Mutabaah ada; Evaluasi/Sekolah/Pengguna tidak ada\n"
        "3. Loop nav ANGGOTA",
        "Nav sesuai ANGGOTA saja",
        "ahmad.fauzi.sman1.g1@gmail.com",
    )
    add(
        "ROLE-06",
        "ROLE",
        "Nested drill-down sekolah→kelompok→anggota",
        "Happy",
        "SUPERADMIN",
        "Seed SMAN 1 punya kelompok & anggota",
        "1. Login Superadmin\n2. Buka Sekolah → klik SMAN 1\n"
        "3. Verifikasi bagian PJ dan daftar Kelompok tampil\n"
        "4. Klik kelompok pertama\n"
        "5. Verifikasi detail (banner gender, pembina, anggota)\n"
        "6. Klik anggota pertama (jika ada)\n7. Verifikasi detail anggota",
        "Drill-down lancar tanpa blank/crash; data seed tampil",
        "SMAN 1",
    )
    add(
        "ROLE-07",
        "ROLE",
        "RoleSwitcher multi-role PEMBINA↔ANGGOTA",
        "Happy",
        "Multi-role",
        "Login multi.role.sman1@gmail.com",
        "1. Login multi-role (default PEMBINA)\n"
        "2. Pastikan menu Evaluasi ada; Mutabaah tidak ada\n"
        "3. Buka RoleSwitcher → pilih ANGGOTA\n"
        "4. Pastikan menu berubah: Mutabaah ada; Evaluasi hilang\n"
        "5. Switch kembali ke PEMBINA",
        "Menu dan konteks berubah sesuai role aktif; tidak perlu logout",
        "multi.role.sman1@gmail.com / !Password123",
    )
    add(
        "ROLE-08",
        "ROLE",
        "ACL smoke URL terlarang → /dashboard",
        "Negatif",
        "ANGGOTA, PEMBINA, ADMIN",
        "Login sebagai role terbatas",
        "Anggota:\n1. Ketik /users → Enter\n2. Ketik /schools/new → Enter\n"
        "3. Ketik /config → Enter\nPembina:\n4. Ketik /users → Enter\n"
        "Admin:\n5. Ketik /users → Enter",
        "RoleGuard mengarahkan ke /dashboard; data sensitif tidak tampil",
        "URL terlarang per role",
    )

    # USERS
    add(
        "USERS-01",
        "USERS",
        "SUPERADMIN lihat daftar pengguna",
        "Happy",
        "SUPERADMIN",
        "Login Superadmin",
        "1. Buka menu Pengguna (/users)\n2. Amati daftar user, filter/search jika ada\n"
        "3. Klik salah satu user ke detail",
        "Daftar user tampil; detail terbuka; role badges terlihat",
        "—",
    )
    add(
        "USERS-02",
        "USERS",
        "ADMIN ditolak akses /users",
        "Negatif",
        "ADMIN",
        "Login Admin",
        "1. Pastikan sidebar tanpa Pengguna\n2. Akses langsung URL /users dan /users/invite",
        "Redirect /dashboard atau akses ditolak; tidak bisa undang admin via UI",
        "—",
    )
    add(
        "USERS-03",
        "USERS",
        "Undang ADMIN baru",
        "Happy",
        "SUPERADMIN",
        "Email unik belum terdaftar / tidak ada undangan pending",
        "1. Buka /users/invite\n2. Isi nama, email unik, role ADMIN\n"
        "3. (Opsional) centang alsoAsPembina jika ada\n4. Submit\n"
        "5. Buka /invitations — pastikan undangan pending muncul",
        "Undangan berhasil dibuat; muncul di list; email/log berisi link aktivasi",
        "email unik mis. admin.uji.p0+timestamp@example.com",
    )
    add(
        "USERS-04",
        "USERS",
        "Undang email yang sudah pending ditolak",
        "Negatif",
        "SUPERADMIN",
        "Sudah ada undangan pending untuk email X",
        "1. Ulangi undang dengan email X yang sama",
        "Error: undangan pending sudah ada; tidak membuat duplikat",
        "Email dari USERS-03",
    )
    add(
        "USERS-05",
        "USERS",
        "Assign & hapus role tambahan",
        "Happy",
        "SUPERADMIN",
        "User existing di detail",
        "1. Buka detail user yang eligible\n"
        "2. Assign role tambahan yang diizinkan CanInvite (mis. PEMBINA)\n"
        "3. Verifikasi role muncul\n4. Hapus role tambahan tersebut\n"
        "5. Verifikasi role hilang",
        "Assign sukses; remove sukses; role utama yang tidak dihapus tetap ada",
        "User non-superadmin",
    )
    add(
        "USERS-06",
        "USERS",
        "Assign role di luar CanInvite ditolak",
        "Negatif",
        "ADMIN",
        "Login sebagai Admin (Skip jika UI assign tidak ada)",
        "1. Coba assign role ADMIN ke user lain\n2. Amati respons",
        "403 / pesan tidak diizinkan; role tidak berubah",
        "—",
    )

    # INVI
    add(
        "INVI-01",
        "INVI",
        "Matrix CanInvite — SUPERADMIN undang role eligible",
        "Happy",
        "SUPERADMIN",
        "Email unik per role",
        "1. Undang berturut-turut: ADMIN (/users/invite), PJ (sekolah), "
        "PEMBINA, ANGGOTA (kelompok)\n"
        "2. Pastikan tiap undangan pending tercatat",
        "Semua undangan sesuai matrix berhasil; role di luar matrix tidak tersedia di UI",
        "Email unik per undangan",
    )
    add(
        "INVI-02",
        "INVI",
        "PJ hanya boleh undang PEMBINA",
        "Happy",
        "PJ_SEKOLAH",
        "Login PJ SMAN 1",
        "1. Dari sekolah/kelompok, buka form undang\n"
        "2. Pastikan opsi role hanya PEMBINA (atau UI default)\n"
        "3. Undang pembina baru dengan gender\n4. Submit",
        "Undangan PEMBINA sukses; tidak bisa pilih ADMIN/PJ",
        "pembina.baru.p0@example.com + gender",
    )
    add(
        "INVI-03",
        "INVI",
        "Pembina undang ANGGOTA ke kelompok",
        "Happy",
        "PEMBINA",
        "Login pembina; punya kelompok",
        "1. Buka detail kelompok\n2. Klik Undang Anggota\n"
        "3. Isi nama, email, gender = gender kelompok\n4. Submit\n"
        "5. Cek list undangan pending",
        "Undangan anggota dibuat; terikat groupId yang benar",
        "Gender harus match kelompok",
    )
    add(
        "INVI-04",
        "INVI",
        "Undang ANGGOTA gender mismatch ditolak",
        "Negatif",
        "PEMBINA",
        "Kelompok Ikhwan (L) atau Akhwat (P)",
        "1. Undang anggota dengan gender berlawanan dari kelompok\n2. Submit",
        "Error 400 / validasi gender; undangan tidak dibuat",
        "Gender ≠ kelompok",
    )
    add(
        "INVI-05",
        "INVI",
        "Gender wajib untuk PJ/PEMBINA/ANGGOTA",
        "Negatif",
        "ADMIN / SUPERADMIN",
        "Form undangan",
        "1. Coba submit undangan PJ/PEMBINA/ANGGOTA tanpa gender\n2. Amati validasi FE/API",
        "Submit diblokir; pesan gender wajib",
        "Gender kosong",
    )
    add(
        "INVI-06",
        "INVI",
        "List undangan: resend & batalkan",
        "Happy",
        "SUPERADMIN",
        "Ada undangan pending milik pengundang",
        "1. Buka /invitations\n2. Pilih undangan pending milik sendiri\n"
        "3. Klik Resend — pastikan sukses\n4. Klik Batal/Cancel\n"
        "5. Pastikan token lama tidak bisa dipakai set-password",
        "Resend sukses; cancel sukses; aktivasi token lama gagal setelah cancel",
        "Undangan pending",
    )
    add(
        "INVI-07",
        "INVI",
        "Resend/batal undangan orang lain ditolak",
        "Negatif",
        "ADMIN / PJ",
        "Ada undangan dibuat Superadmin (bukan milik penguji)",
        "1. Login sebagai Admin/PJ\n"
        "2. Coba resend/batal undangan yang bukan miliknya (jika terlihat)",
        "Aksi ditolak / undangan tidak muncul di list milik penguji",
        "—",
    )
    add(
        "INVI-08",
        "INVI",
        "Aktivasi undangan → set-password → role + relasi",
        "Happy",
        "Public",
        "Undangan pending anggota/pembina baru",
        "1. Ambil link dari email/log\n2. Set password\n3. Login\n"
        "4. Verifikasi role aktif\n"
        "5. Anggota: pastikan masuk GroupMember\n"
        "6. Pembina: pastikan UserSchool ter-link",
        "Akun aktif dengan role & relasi sekolah/kelompok benar",
        "Token undangan baru",
    )

    # SCHOOLS
    add(
        "SCH-01",
        "SCHOOLS",
        "List & detail sekolah SUPERADMIN/ADMIN",
        "Happy",
        "SUPERADMIN, ADMIN",
        "Login SA atau Admin",
        "1. Buka /schools\n2. Pastikan SMAN 1 ada di list\n"
        "3. Buka detail\n4. Verifikasi PJ, kelompok, info sekolah",
        "List & detail lengkap; tidak error",
        "SMAN 1",
    )
    add(
        "SCH-02",
        "SCHOOLS",
        "PJ hanya melihat sekolah ter-assign",
        "Happy",
        "PJ_SEKOLAH",
        "Login PJ SMAN 1",
        "1. Buka /schools\n2. Pastikan hanya sekolah assign (SMAN 1)\n"
        "3. Coba akses schoolId lain via URL jika diketahui",
        "Hanya sekolah sendiri; schoolId lain → 403 / tidak tampil data",
        "SMAN 1 vs school lain",
    )
    add(
        "SCH-03",
        "SCHOOLS",
        "Buat sekolah baru + PJ (undangan)",
        "Happy",
        "SUPERADMIN atau ADMIN",
        "Login SA/Admin",
        "1. Buka /schools/new\n2. Isi nama sekolah unik\n"
        "3. Isi data PJ (nama, email unik, gender, HP jika ada)\n"
        "4. Pilih mode undangan\n5. Submit\n"
        "6. Verifikasi sekolah di list & undangan PJ pending",
        "Sekolah dibuat; undangan PJ terbuat",
        "Nama sekolah unik + email PJ unik + gender",
    )
    add(
        "SCH-04",
        "SCHOOLS",
        "Buat sekolah + PJ set password langsung",
        "Happy",
        "SUPERADMIN atau ADMIN",
        "Login SA/Admin",
        "1. /schools/new\n2. Isi sekolah + PJ\n"
        "3. Pilih set password langsung (jika opsi ada)\n"
        "4. Isi password PJ\n5. Submit\n6. Logout → login sebagai PJ baru",
        "Sekolah & PJ aktif langsung; PJ bisa login dan melihat sekolahnya",
        "Password PJ sesuai aturan app",
    )
    add(
        "SCH-05",
        "SCHOOLS",
        "Validasi wajib nama sekolah & data PJ",
        "Negatif",
        "ADMIN",
        "/schools/new",
        "1. Submit form kosong\n2. Isi nama sekolah saja tanpa PJ\n"
        "3. Isi email PJ invalid",
        "Validasi FE/API menolak; sekolah tidak terbuat",
        "Input invalid",
    )
    add(
        "SCH-06",
        "SCHOOLS",
        "Undang PJ tambahan",
        "Happy",
        "SUPERADMIN / ADMIN",
        "Sekolah existing",
        "1. Dari detail sekolah → Undang PJ\n2. Isi data PJ baru + gender\n"
        "3. Submit\n4. Cek undangan / daftar PJ",
        "PJ tambahan terundang; muncul di detail setelah aktivasi",
        "Email PJ baru",
    )
    add(
        "SCH-07",
        "SCHOOLS",
        "Ganti / hapus PJ",
        "Happy",
        "SUPERADMIN / ADMIN",
        "Sekolah dengan ≥1 PJ (gunakan sekolah uji, bukan seed kritis)",
        "1. Dari detail → Ganti PJ\n2. Atau hapus PJ non-kritis\n"
        "3. Verifikasi daftar PJ ter-update\n"
        "4. Jika ada aturan tidak boleh hapus PJ terakhir, uji dan catat",
        "Perubahan PJ tersimpan; aturan bisnis PJ terakhir dihormati",
        "Sekolah uji dari SCH-03",
    )
    add(
        "SCH-08",
        "SCHOOLS",
        "PEMBINA/ANGGOTA tidak akses CRUD sekolah",
        "Negatif",
        "PEMBINA, ANGGOTA",
        "Login role terkait",
        "1. Pastikan menu Sekolah tidak ada\n2. Akses /schools dan /schools/new via URL",
        "Ditolak / redirect dashboard; tidak bisa create",
        "—",
    )

    # KELOMPOK
    add(
        "GRP-01",
        "KELOMPOK",
        "PJ buat kelompok dengan pembina existing",
        "Happy",
        "PJ_SEKOLAH",
        "Login PJ; ada pembina di sekolah",
        "1. Buka detail SMAN 1 → Kelompok Baru\n"
        "2. Isi nama (≥2 karakter), level (LEVEL_1/LEVEL_2), gender\n"
        "3. Pilih pembina existing (jangan isi pembina baru bersamaan)\n"
        "4. Submit\n5. Buka detail kelompok baru",
        "Kelompok dibuat; pembina ter-assign; banner gender sesuai",
        "Nama unik, level, gender, pembina existing",
    )
    add(
        "GRP-02",
        "KELOMPOK",
        "Buat kelompok + undang pembina baru",
        "Happy",
        "ADMIN / PJ",
        "Email pembina baru unik",
        "1. Kelompok Baru\n2. Isi meta kelompok\n"
        "3. Pilih mode pembina baru (undangan atau password)\n"
        "4. Isi data pembina + gender\n"
        "5. Jangan sekaligus pilih pembina existing\n6. Submit",
        "Kelompok + undangan/akun pembina terbuat; UserSchool ter-link setelah aktivasi",
        "Pembina baru saja",
    )
    add(
        "GRP-03",
        "KELOMPOK",
        "Tolak existing+baru pembina sekaligus / nama pendek",
        "Negatif",
        "PJ_SEKOLAH",
        "Form kelompok baru",
        "1. Coba submit dengan pembina existing DAN data pembina baru sekaligus\n"
        "2. Coba nama kelompok 1 karakter",
        "400 / validasi: tidak boleh keduanya; nama minimal 2 karakter",
        "Input invalid",
    )
    add(
        "GRP-04",
        "KELOMPOK",
        "Detail kelompok lengkap",
        "Smoke",
        "SUPERADMIN / PEMBINA",
        "Kelompok seed ada",
        "1. Buka /kelompok/:id\n"
        "2. Cek banner gender, nama, level, pembina, daftar anggota",
        "Semua section utama tampil; tidak blank",
        "Kelompok SMAN 1",
    )
    add(
        "GRP-05",
        "KELOMPOK",
        "Edit kelompok & ganti pembina (PJ/Admin)",
        "Happy",
        "PJ_SEKOLAH / ADMIN",
        "Kelompok uji",
        "1. Buka /kelompok/:id/edit\n2. Ubah nama/level\n"
        "3. Ganti pembina ke pembina lain di sekolah\n4. Simpan\n5. Verifikasi detail",
        "Perubahan tersimpan; pembina baru tampil di detail",
        "Pembina pengganti di sekolah sama",
    )
    add(
        "GRP-06",
        "KELOMPOK",
        "PEMBINA tidak boleh ganti pembina kelompok",
        "Negatif",
        "PEMBINA",
        "Login pembina",
        "1. Buka detail kelompoknya\n2. Coba akses /kelompok/:id/edit\n"
        "3. Jika form terbuka, coba ganti pembina",
        "Edit dibatasi / 403; pembina tidak berubah",
        "—",
    )
    add(
        "GRP-07",
        "KELOMPOK",
        "Undang anggota → aktivasi masuk GroupMember",
        "Happy",
        "PEMBINA",
        "Undangan anggota pending dari INVI-03 atau buat baru",
        "1. Selesaikan set-password anggota\n2. Login sebagai anggota baru\n"
        "3. Sebagai pembina, buka detail kelompok — pastikan anggota muncul\n"
        "4. Buka detail anggota",
        "Anggota terdaftar di kelompok; detail anggota bisa dibuka",
        "Anggota baru hasil undangan",
    )
    add(
        "GRP-08",
        "KELOMPOK",
        "Edit profil anggota + validasi gender kelompok",
        "Happy",
        "PEMBINA / PJ",
        "Ada anggota di kelompok",
        "1. Buka edit anggota\n2. Ubah nama/HP valid → simpan (sukses)\n"
        "3. Coba ubah gender melanggar gender kelompok → simpan",
        "Edit valid sukses; gender mismatch ditolak 400",
        "Anggota seed atau uji",
    )

    # EVAL
    add(
        "EVAL-01",
        "EVAL",
        "PEMBINA buka list evaluasi",
        "Happy",
        "PEMBINA",
        "Login pembina",
        "1. Buka /evaluasi dari sidebar\n2. Scroll list (infinite) jika banyak data\n"
        "3. Pastikan item evaluasi / empty state jelas",
        "Halaman load; list atau empty state; tidak 403 untuk pembina pemilik kelompok",
        "—",
    )
    add(
        "EVAL-02",
        "EVAL",
        "ANGGOTA tidak akses evaluasi",
        "Negatif",
        "ANGGOTA",
        "Login anggota",
        "1. Pastikan menu Evaluasi tidak ada\n2. Akses URL /evaluasi dan /evaluasi/isi",
        "Tidak di menu; API/UI menolak atau tidak menampilkan data pembina",
        "—",
    )
    add(
        "EVAL-03",
        "EVAL",
        "Isi evaluasi pekan — simpan draft",
        "Happy",
        "PEMBINA",
        "Pembina punya kelompok; pilih minggu yang belum ada evaluasi",
        "1. Buka /evaluasi/isi (atau tombol Isi evaluasi)\n"
        "2. Pilih kelompok & tanggal minggu (weekDate)\n"
        "3. Isi kehadiran anggota, catatan\n4. Simpan draft\n"
        "5. Kembali ke list — pastikan draft muncul",
        "Draft tersimpan; bisa dibuka lagi di detail/edit",
        "weekDate unik untuk group",
    )
    add(
        "EVAL-04",
        "EVAL",
        "Submit evaluasi (+ foto opsional)",
        "Happy",
        "PEMBINA",
        "Ada draft atau form isi siap submit",
        "1. Buka evaluasi draft\n2. Lengkapi field wajib\n"
        "3. Upload foto jika fitur ada (opsional)\n4. Klik Submit\n"
        "5. Verifikasi status submitted\n"
        "6. Cek poin pembina jika ditampilkan (tepat waktu 10 / terlambat 5)",
        "Status submitted; foto tampil jika diunggah; poin wajar jika UI menampilkan",
        "Foto JPG/PNG kecil < limit",
    )
    add(
        "EVAL-05",
        "EVAL",
        "Duplikat evaluasi pekan yang sama → 409",
        "Negatif",
        "PEMBINA",
        "Sudah ada evaluasi untuk (groupId, weekDate)",
        "1. Coba Isi evaluasi lagi untuk kelompok & minggu yang sama\n2. Submit create",
        'Ditolak 409 / pesan "sudah ada"; tidak membuat duplikat; diarahkan edit existing',
        "Pekan dari EVAL-03/04",
    )
    add(
        "EVAL-06",
        "EVAL",
        "Edit evaluasi via detail (PUT)",
        "Happy",
        "PEMBINA",
        "Evaluasi existing milik pembina",
        "1. Buka /evaluasi/:id\n2. Edit catatan/kehadiran\n3. Simpan\n4. Refresh",
        "Perubahan tersimpan; tidak create baris baru",
        "—",
    )
    add(
        "EVAL-07",
        "EVAL",
        "Pembina tidak akses evaluasi kelompok lain",
        "Negatif",
        "PEMBINA",
        "Ketahu evaluationId milik pembina lain",
        "1. Akses /evaluasi/:id milik pembina lain via URL\n2. Coba edit/submit",
        "403 / tidak ditemukan; data tidak bocor",
        "ID asing",
    )
    add(
        "EVAL-08",
        "EVAL",
        "PJ scoped melihat evaluasi sekolahnya",
        "Smoke",
        "PJ_SEKOLAH",
        "Login PJ; ada evaluasi di sekolah",
        "1. Dari UI terkait sekolah/kelompok/evaluasi\n"
        "2. Pastikan bisa melihat ringkasan evaluasi kelompok di sekolahnya\n"
        "3. Pastikan tidak melihat data sekolah lain",
        "Data terbatas scope sekolah PJ",
        "SMAN 1",
    )

    add(
        "PROF-01",
        "PROFILE",
        "Buka & update profil sendiri",
        "Smoke",
        "Semua",
        "Login salah satu role",
        "1. Buka /profile\n2. Verifikasi data nama/email tampil\n"
        "3. Ubah field yang diizinkan lalu simpan\n4. Refresh",
        "Profil load; update sukses untuk field allowed",
        "Jangan ubah email ke konflik jika tidak perlu",
    )
    add(
        "DASH-01",
        "DASHBOARD",
        "Dashboard load per role",
        "Smoke",
        "Semua role",
        "Login tiap role",
        "1. Setelah login, amati /dashboard\n2. Pastikan widget tidak crash\n3. Tidak blank page",
        "Dashboard render sesuai role",
        "—",
    )
    add(
        "E2E-01",
        "E2E",
        "Journey kritis: undang → aktivasi → kelompok → evaluasi",
        "Happy",
        "ADMIN → PEMBINA → ANGGOTA",
        "Env cukup untuk data uji baru; catat semua email yang dibuat",
        "1. Admin pastikan/buat sekolah uji\n"
        "2. PJ/Admin buat kelompok baru + pembina (undangan)\n"
        "3. Aktivasi pembina (set-password)\n"
        "4. Login pembina → undang anggota (gender match)\n"
        "5. Aktivasi anggota\n"
        "6. Login pembina → isi & submit evaluasi minggu ini\n"
        "7. Verifikasi anggota di kelompok & evaluasi tersimpan",
        "Seluruh rantai onboarding → evaluasi sukses; data konsisten di UI",
        "Gunakan email unik bertimestamp",
    )
    return cases


def main() -> None:
    wb = Workbook()
    cases = build_cases()
    n = len(cases)
    end_row = 4 + n

    # ----- 01 Petunjuk -----
    ws = wb.active
    ws.title = "01_Petunjuk"
    ws["A1"] = "AISI — Manual Test Plan (P0 Only)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws["A2"] = (
        f"Dibuat: {date.today().isoformat()} | Scope: Critical / Smoke path saja | "
        "Stack: Vite (:5173) + Go API (:4000)"
    )
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="5D6D7E")
    ws.merge_cells("A2:B2")

    petunjuk = [
        (
            "Tujuan",
            "Dokumen ini adalah rencana uji manual komprehensif untuk case prioritas P0 "
            "(critical smoke + happy path inti + failure kritis) pada aplikasi AISI.",
        ),
        (
            "Cakupan P0",
            "AUTH, ROLE/NAV/ACL, USERS, INVITATIONS, SCHOOLS, KELOMPOK, EVALUATIONS. "
            "Modul Agenda, Mutabaah, IC, KKS, Analitik, Materi, Config detail → P1 "
            "(kecuali smoke buka halaman config SUPERADMIN di nav loop).",
        ),
        (
            "Prasyarat lingkungan",
            "1) Postgres jalan + pnpm db:deploy\n"
            "2) Seed demo: pnpm db:seed\n"
            "3) API Go di :4000\n"
            "4) Frontend Vite di :5173\n"
            "5) Browser Chrome/Edge; gunakan Incognito per role",
        ),
        (
            "Cara memakai sheet",
            "1) Baca 02_Akun_Seed untuk kredensial\n"
            "2) Lihat 03_Matriks_Akses untuk ekspektasi menu/role\n"
            "3) Eksekusi case di 04_Test_Cases sesuai urutan Case ID\n"
            "4) Isi Status, Actual Result, Tester, Tanggal\n"
            "5) Pantau ringkasan di 05_Ringkasan\n"
            "6) Smoke cepat: 06_Checklist_Cepat\n"
            "7) Catat bug di 07_Bug_Log",
        ),
        (
            "Konvensi Status",
            "Not Run | Pass | Fail | Blocked | Skip",
        ),
        (
            "Konvensi Tipe",
            "Happy = alur sukses utama | Negatif = validasi/error/ACL | "
            "Smoke = cek buka halaman & tidak blank/crash",
        ),
        (
            "Urutan eksekusi",
            "AUTH → ROLE → USERS → INVI → SCHOOLS → KELOMPOK → EVAL → E2E",
        ),
        (
            "Catatan undangan email",
            "Di lokal, email Resend bisa kosong — cek log console API untuk link "
            "set-password / accept-role. Token seed ada di sheet Akun.",
        ),
        (
            "Referensi repo",
            "docs/E2E_MATRIX.md, docs/STAGING.md, "
            "apps/web-vite/e2e/fixtures/users.ts, App.tsx routes",
        ),
        (
            "Kriteria hijau P0",
            "Semua case P0 = Pass (atau Skip beralasan). "
            "Tidak ada blank page / console pageerror kritis pada smoke nav 5 role. "
            "E2E-01 Pass sebelum cutover staging.",
        ),
    ]
    row = 4
    for label, text in petunjuk:
        ws.cell(row=row, column=1, value=label).font = Font(
            bold=True, name="Calibri", size=11, color="1B4F72"
        )
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).alignment = wrap
        ws.cell(row=row, column=1).border = thin
        ws.cell(row=row, column=2, value=text).alignment = wrap
        ws.cell(row=row, column=2).border = thin
        ws.cell(row=row, column=2).font = Font(name="Calibri", size=10)
        ws.row_dimensions[row].height = max(45, 15 * (text.count("\n") + 2))
        row += 1
    set_widths(ws, [28, 110])
    ws.freeze_panes = "A4"

    # ----- 02 Akun -----
    ws2 = wb.create_sheet("02_Akun_Seed")
    ws2["A1"] = "Akun Seed untuk Uji Manual (dari packages/shared/prisma/seed.ts)"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:F1")
    ws2["A2"] = (
        "Peringatan: password seed hanya untuk lokal/demo. "
        "Jangan pakai di production tanpa rotasi."
    )
    ws2["A2"].font = Font(name="Calibri", size=10, italic=True, color="C0392B")

    headers2 = [
        "Role",
        "Label",
        "Email",
        "Password",
        "Catatan Penggunaan P0",
        "Primary Active Role",
    ]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=i, value=h)
    style_header(ws2, 4, len(headers2))

    accounts = [
        (
            "SUPERADMIN",
            "Superadmin",
            "fuadinaqi@gmail.com",
            "!Superadmin123",
            "Users, undang ADMIN, CRUD sekolah/PJ, undangan, assign/hapus role, nested, config smoke",
            "SUPERADMIN",
        ),
        (
            "ADMIN",
            "Admin",
            "fuadiproject@gmail.com",
            "!Admin123",
            "Sekolah, undang PJ/pembina/anggota, tidak akses /users",
            "ADMIN",
        ),
        (
            "PJ_SEKOLAH",
            "PJ Sekolah",
            "usamah_sman1@gmail.com",
            "!Password123",
            "Scope SMAN 1; buat kelompok; undang pembina; /pembina",
            "PJ_SEKOLAH",
        ),
        (
            "PEMBINA",
            "Pembina",
            "budi.santoso.sman1.pembina@gmail.com",
            "!Password123",
            "Evaluasi mingguan, undang anggota, detail kelompok/anggota",
            "PEMBINA",
        ),
        (
            "ANGGOTA",
            "Anggota",
            "ahmad.fauzi.sman1.g1@gmail.com",
            "!Password123",
            "Login, dashboard, profil; bukan evaluasi/sekolah CRUD",
            "ANGGOTA",
        ),
        (
            "PEMBINA+ANGGOTA",
            "Multi-role",
            "multi.role.sman1@gmail.com",
            "!Password123",
            "Uji RoleSwitcher; default prioritas PEMBINA > ANGGOTA",
            "PEMBINA (default)",
        ),
    ]
    for i, a in enumerate(accounts):
        r = 5 + i
        for c, v in enumerate(a, 1):
            ws2.cell(row=r, column=c, value=v)
        style_data_row(ws2, r, 6)
        ws2.row_dimensions[r].height = 35

    ws2["A12"] = "Token Seed & Data Referensi"
    ws2["A12"].font = subtitle_font
    for i, h in enumerate(["Nama", "Nilai", "Cara Pakai"], 1):
        ws2.cell(row=13, column=i, value=h)
    style_header(ws2, 13, 3)
    tokens = [
        (
            "SEED_INVITE_TOKEN",
            "00000000-0000-4000-8000-000000000001",
            "Buka /set-password?token=<nilai> untuk user undangan baru (jika masih valid)",
        ),
        (
            "SEED_ACCEPT_ROLE_TOKEN",
            "00000000-0000-4000-8000-000000000002",
            "Buka /accept-role?token=<nilai> untuk existing user menerima role tambahan",
        ),
        ("SEED_SCHOOL_NAME", "SMAN 1", "Sekolah utama untuk drill-down nested & scope PJ"),
        ("Invitation expiry", "7 hari", "Undangan kadaluarsa setelah 7 hari dari dibuat"),
        (
            "Role priority (multi)",
            "SUPERADMIN > ADMIN > PJ_SEKOLAH > PEMBINA > ANGGOTA",
            "Menentukan activeRole default saat login",
        ),
    ]
    for i, t in enumerate(tokens):
        r = 14 + i
        for c, v in enumerate(t, 1):
            ws2.cell(row=r, column=c, value=v)
        style_data_row(ws2, r, 3)

    ws2["A21"] = "Matrix REST CanInvite (aturan undangan)"
    ws2["A21"].font = subtitle_font
    for i, h in enumerate(["Pengundang", "Boleh mengundang role"], 1):
        ws2.cell(row=22, column=i, value=h)
    style_header(ws2, 22, 2)
    for i, t in enumerate(
        [
            ("SUPERADMIN", "ADMIN, PJ_SEKOLAH, PEMBINA, ANGGOTA"),
            ("ADMIN", "PJ_SEKOLAH, PEMBINA, ANGGOTA"),
            ("PJ_SEKOLAH", "PEMBINA"),
            ("PEMBINA", "ANGGOTA"),
            ("ANGGOTA", "— (tidak boleh mengundang)"),
        ]
    ):
        r = 23 + i
        for c, v in enumerate(t, 1):
            ws2.cell(row=r, column=c, value=v)
        style_data_row(ws2, r, 2)

    set_widths(ws2, [20, 22, 42, 18, 70, 22])
    ws2.freeze_panes = "A5"

    # ----- 03 Matriks -----
    ws3 = wb.create_sheet("03_Matriks_Akses")
    ws3["A1"] = "Matriks Akses UI P0 (ekspektasi menu / fitur per role)"
    ws3["A1"].font = title_font
    ws3.merge_cells("A1:G1")
    ws3["A2"] = (
        "Ya = boleh akses via UI | Scope = hanya data ter-assign | "
        "— = tidak di menu / RoleGuard redirect | API 403 = UI mungkin terbuka tapi API menolak"
    )
    ws3["A2"].font = Font(name="Calibri", size=9, italic=True)

    mh = [
        "Halaman / Fitur",
        "SUPERADMIN",
        "ADMIN",
        "PJ_SEKOLAH",
        "PEMBINA",
        "ANGGOTA",
        "Catatan Uji",
    ]
    for i, h in enumerate(mh, 1):
        ws3.cell(row=4, column=i, value=h)
    style_header(ws3, 4, len(mh))

    matrix = [
        ("/dashboard", "Ya", "Ya", "Ya", "Ya", "Ya", "Smoke: konten tidak blank"),
        ("/profile", "Ya", "Ya", "Ya", "Ya", "Ya", "Edit profil dasar"),
        ("/notifications", "Ya", "Ya", "Ya", "Ya", "Ya", "Smoke buka halaman"),
        ("/users, /users/invite", "Ya", "—", "—", "—", "—", "Hanya SUPERADMIN di FE"),
        ("/invitations", "Ya", "Ya*", "Ya*", "Ya*", "—", "RoleGuard: SA/Admin/PJ/Pembina"),
        ("/schools list & detail", "Ya", "Ya", "Scope", "—", "—", "PJ hanya sekolah assign"),
        ("/schools/new", "Ya", "Ya", "—", "—", "—", "Buat sekolah + PJ"),
        ("Undang/Ganti/Hapus PJ", "Ya", "Ya", "—", "—", "—", "SCH-06, SCH-07"),
        ("Buat/Edit kelompok", "Ya", "Ya", "Ya", "—", "—", "/schools/:id/kelompok/baru"),
        ("Undang/Edit anggota", "Ya", "Ya", "Ya", "Ya", "—", "Gender wajib match kelompok"),
        ("/evaluasi*", "—**", "—**", "Scope baca", "Ya", "—", "Tidak di nav SA/Admin"),
        ("/pembina", "—", "—", "Ya", "—", "—", "Daftar pembina sekolah"),
        ("/config (smoke)", "Ya", "Ya", "IC only", "IC only", "—", "P0: smoke buka untuk SA"),
        ("RoleSwitcher", "jika multi", "jika multi", "jika multi", "jika multi", "jika multi", "Akun multi.role"),
    ]
    for i, m in enumerate(matrix):
        r = 5 + i
        for c, v in enumerate(m, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = center if 1 < c < 7 else wrap
            cell.font = Font(name="Calibri", size=10)
            if c in (2, 3, 4, 5, 6):
                if v == "Ya" or str(v).startswith("Ya"):
                    cell.fill = yes_fill
                elif v == "—":
                    cell.fill = no_fill
                elif any(x in str(v) for x in ("Scope", "IC", "multi", "baca")):
                    cell.fill = scope_fill

    ws3["A22"] = "Nav item yang HARUS muncul (smoke ROLE)"
    ws3["A22"].font = subtitle_font
    for i, h in enumerate(["Role", "Menu yang harus ada (urut sidebar)"], 1):
        ws3.cell(row=23, column=i, value=h)
    style_header(ws3, 23, 2)
    navs = [
        (
            "SUPERADMIN",
            "Beranda, Sekolah, Pengguna, Undangan, Agenda, Materi, Analitik, KKS, Pengaturan, Notifikasi, Profil",
        ),
        (
            "ADMIN",
            "Beranda, Sekolah, Agenda, Materi, Analitik, KKS, Pengaturan, Notifikasi, Profil",
        ),
        (
            "PJ_SEKOLAH",
            "Beranda, Sekolah, Pembina, Indikator Capaian, Agenda, Materi, Analitik, KKS, Notifikasi, Profil",
        ),
        (
            "PEMBINA",
            "Beranda, Evaluasi, Indikator Capaian, Agenda, Materi, KKS, Notifikasi, Profil",
        ),
        ("ANGGOTA", "Beranda, Mutabaah, Agenda, KKS, Notifikasi, Profil"),
    ]
    for i, nav in enumerate(navs):
        r = 24 + i
        for c, v in enumerate(nav, 1):
            ws3.cell(row=r, column=c, value=v)
        style_data_row(ws3, r, 2)
        ws3.row_dimensions[r].height = 30

    set_widths(ws3, [28, 14, 14, 14, 14, 14, 40])
    ws3.freeze_panes = "A5"

    # ----- 04 Test Cases -----
    ws4 = wb.create_sheet("04_Test_Cases")
    ws4["A1"] = "Test Cases Manual P0 — AISI"
    ws4["A1"].font = title_font
    ws4.merge_cells("A1:N1")
    ws4["A2"] = (
        "Isi kolom Status / Actual Result / Tester / Tanggal saat eksekusi. "
        "Filter kolom Modul atau Role untuk fokus sesi uji."
    )
    ws4["A2"].font = Font(name="Calibri", size=9, italic=True)

    cols = [
        "Case ID",
        "Modul",
        "Judul Case",
        "Tipe",
        "Prioritas",
        "Role Penguji",
        "Prekondisi",
        "Langkah Uji (Cara Test)",
        "Hasil Diharapkan",
        "Data Uji",
        "Status",
        "Actual Result",
        "Tester",
        "Tanggal",
    ]
    for i, h in enumerate(cols, 1):
        ws4.cell(row=4, column=i, value=h)
    style_header(ws4, 4, len(cols))
    ws4.row_dimensions[4].height = 30
    ws4.auto_filter.ref = f"A4:N{end_row}"
    ws4.freeze_panes = "A5"

    for i, case in enumerate(cases):
        r = 5 + i
        for c, v in enumerate(case, 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = wrap
            cell.font = Font(name="Calibri", size=9)
        tipe = case[3]
        if tipe == "Happy":
            ws4.cell(row=r, column=4).fill = happy_fill
        elif tipe == "Negatif":
            ws4.cell(row=r, column=4).fill = neg_fill
        else:
            ws4.cell(row=r, column=4).fill = smoke_fill
        ws4.cell(row=r, column=5).fill = p0_fill
        ws4.cell(row=r, column=5).alignment = center
        ws4.cell(row=r, column=1).alignment = center
        ws4.cell(row=r, column=11).alignment = center
        steps_lines = case[7].count("\n") + 1
        ws4.row_dimensions[r].height = min(160, max(50, steps_lines * 12 + 20))

    dv = DataValidation(
        type="list",
        formula1='"Not Run,Pass,Fail,Blocked,Skip"',
        allow_blank=True,
    )
    ws4.add_data_validation(dv)
    dv.add(f"K5:K{end_row}")
    ws4.conditional_formatting.add(
        f"K5:K{end_row}", FormulaRule(formula=['$K5="Pass"'], fill=pass_fill)
    )
    ws4.conditional_formatting.add(
        f"K5:K{end_row}", FormulaRule(formula=['$K5="Fail"'], fill=fail_fill)
    )
    ws4.conditional_formatting.add(
        f"K5:K{end_row}", FormulaRule(formula=['$K5="Blocked"'], fill=blocked_fill)
    )
    set_widths(ws4, [10, 12, 42, 10, 10, 22, 36, 55, 42, 28, 12, 28, 14, 12])

    # ----- 05 Ringkasan -----
    ws5 = wb.create_sheet("05_Ringkasan")
    ws5["A1"] = "Ringkasan Eksekusi Test P0"
    ws5["A1"].font = title_font
    ws5.merge_cells("A1:B1")
    ws5["A3"] = "Metrik"
    ws5["B3"] = "Nilai"
    style_header(ws5, 3, 2)

    metrics = [
        (4, "Total Case P0", f"=COUNTA('04_Test_Cases'!A5:A{end_row})"),
        (5, "Not Run", f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Not Run\")"),
        (6, "Pass", f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Pass\")"),
        (7, "Fail", f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Fail\")"),
        (8, "Blocked", f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Blocked\")"),
        (9, "Skip", f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Skip\")"),
        (10, "% Pass (dari dieksekusi excl Not Run/Skip)", "=IF((B6+B7+B8)=0,\"-\",B6/(B6+B7+B8))"),
        (
            11,
            "Status keseluruhan",
            '=IF(AND(B5=0,B7=0,B8=0),"YA — P0 hijau","BELUM — ada Not Run/Fail/Blocked")',
        ),
    ]
    for r, label, formula in metrics:
        ws5.cell(row=r, column=1, value=label)
        ws5.cell(row=r, column=2, value=formula)
        for c in range(1, 3):
            ws5.cell(row=r, column=c).border = thin
            ws5.cell(row=r, column=c).alignment = wrap
        ws5.cell(row=r, column=1).fill = section_fill
        ws5.cell(row=r, column=1).font = Font(bold=True, name="Calibri", size=11)
        ws5.cell(row=r, column=2).font = Font(name="Calibri", size=11)
    ws5["B10"].number_format = "0.0%"
    ws5["B6"].fill = pass_fill
    ws5["B7"].fill = fail_fill
    ws5["B8"].fill = blocked_fill

    ws5["A13"] = "Breakdown per Modul"
    ws5["A13"].font = subtitle_font
    for i, h in enumerate(
        ["Modul", "Jumlah Case", "Pass", "Fail", "Not Run", "Catatan Sesi"], 1
    ):
        ws5.cell(row=14, column=i, value=h)
    style_header(ws5, 14, 6)

    cnt = Counter(c[1] for c in cases)
    moduls = [
        "AUTH",
        "ROLE",
        "USERS",
        "INVI",
        "SCHOOLS",
        "KELOMPOK",
        "EVAL",
        "PROFILE",
        "DASHBOARD",
        "E2E",
    ]
    for i, m in enumerate(moduls):
        r = 15 + i
        ws5.cell(row=r, column=1, value=m)
        ws5.cell(row=r, column=2, value=cnt.get(m, 0))
        ws5.cell(
            row=r,
            column=3,
            value=f"=COUNTIFS('04_Test_Cases'!B:B,A{r},'04_Test_Cases'!K:K,\"Pass\")",
        )
        ws5.cell(
            row=r,
            column=4,
            value=f"=COUNTIFS('04_Test_Cases'!B:B,A{r},'04_Test_Cases'!K:K,\"Fail\")",
        )
        ws5.cell(
            row=r,
            column=5,
            value=f"=COUNTIFS('04_Test_Cases'!B:B,A{r},'04_Test_Cases'!K:K,\"Not Run\")",
        )
        style_data_row(ws5, r, 6)

    ws5["A27"] = "Log Sesi Uji"
    ws5["A27"].font = subtitle_font
    for i, h in enumerate(
        ["Tanggal", "Tester", "Environment (local/staging)", "Build/Commit", "Hasil Ringkas", "Bug terkait"],
        1,
    ):
        ws5.cell(row=28, column=i, value=h)
    style_header(ws5, 28, 6)
    for r in range(29, 34):
        for c in range(1, 7):
            ws5.cell(row=r, column=c).border = thin

    ws5["A36"] = "Definisi Kelulusan P0"
    ws5["A36"].font = subtitle_font
    ws5["A37"] = (
        "1. Semua case Happy & Smoke = Pass\n"
        "2. Case Negatif kritis (login gagal, ACL, 409 evaluasi, gender mismatch) = Pass\n"
        "3. Tidak ada blank page pada nav 5 role\n"
        "4. Journey E2E-01 = Pass sebelum cutover staging\n"
        "5. Fail yang tersisa harus punya ticket bug sebelum release"
    )
    ws5["A37"].alignment = wrap
    ws5.merge_cells("A37:F37")
    ws5.row_dimensions[37].height = 80
    set_widths(ws5, [42, 14, 14, 14, 14, 40])

    # ----- 06 Checklist -----
    ws6 = wb.create_sheet("06_Checklist_Cepat")
    ws6["A1"] = "Checklist Cepat P0 (1 halaman — sesi smoke 30–45 menit)"
    ws6["A1"].font = title_font
    ws6.merge_cells("A1:D1")
    for i, h in enumerate(["#", "Checklist", "Role", "Done (✓)"], 1):
        ws6.cell(row=3, column=i, value=h)
    style_header(ws6, 3, 4)

    checklist = [
        ("Login 5 role seed sukses", "Semua"),
        ("Logout + proteksi /dashboard", "SA"),
        ("Nav tidak blank — Superadmin", "SA"),
        ("Nav tidak blank — Admin", "Admin"),
        ("Nav tidak blank — PJ", "PJ"),
        ("Nav tidak blank — Pembina (+Evaluasi)", "Pembina"),
        ("Nav tidak blank — Anggota (+Mutabaah)", "Anggota"),
        ("Drill-down SMAN 1 → kelompok → anggota", "SA"),
        ("RoleSwitcher multi-role bekerja", "Multi"),
        ("ACL: Anggota tidak masuk /users", "Anggota"),
        ("Undang admin / lihat /invitations", "SA"),
        ("Buat sekolah atau kelompok (satu write flow)", "Admin/PJ"),
        ("Undang anggota gender match", "Pembina"),
        ("Isi + submit evaluasi minggu baru", "Pembina"),
        ("Duplikat evaluasi ditolak", "Pembina"),
        ("Profil & dashboard load", "Semua"),
    ]
    for i, (text, role) in enumerate(checklist, 1):
        r = 3 + i
        ws6.cell(row=r, column=1, value=i)
        ws6.cell(row=r, column=2, value=text)
        ws6.cell(row=r, column=3, value=role)
        ws6.cell(row=r, column=4, value="")
        style_data_row(ws6, r, 4)
        ws6.cell(row=r, column=1).alignment = center
        ws6.cell(row=r, column=4).alignment = center

    dv2 = DataValidation(type="list", formula1='"☐,✓,✗,—"', allow_blank=True)
    ws6.add_data_validation(dv2)
    dv2.add(f"D4:D{3 + len(checklist)}")
    set_widths(ws6, [6, 55, 14, 12])
    ws6.freeze_panes = "A4"

    # ----- 07 Bug Log -----
    ws7 = wb.create_sheet("07_Bug_Log")
    ws7["A1"] = "Bug Log hasil uji manual P0"
    ws7["A1"].font = title_font
    bug_headers = [
        "Bug ID",
        "Case ID terkait",
        "Severity (S0-S3)",
        "Judul",
        "Langkah Reproduce",
        "Expected",
        "Actual",
        "Status",
        "Link Issue",
    ]
    for i, h in enumerate(bug_headers, 1):
        ws7.cell(row=3, column=i, value=h)
    style_header(ws7, 3, len(bug_headers))
    for r in range(4, 20):
        for c in range(1, len(bug_headers) + 1):
            ws7.cell(row=r, column=c).border = thin
            ws7.cell(row=r, column=c).alignment = wrap
    set_widths(ws7, [10, 14, 14, 28, 40, 28, 28, 12, 28])
    ws7["A22"] = (
        "Severity: S0=blocker production | S1=fitur P0 rusak | S2=workaround ada | S3=kosmetik"
    )
    ws7["A22"].font = Font(italic=True, size=9, color="5D6D7E")

    wb.save(OUT)
    print(f"Saved: {OUT}")
    print(f"Total cases: {n}")
    print("Modules:", dict(cnt))


if __name__ == "__main__":
    main()
