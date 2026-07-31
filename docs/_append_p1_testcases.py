"""Append P1 module test cases (Agenda, Materi, Analitik, etc.) to the Excel test plan."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

OUT = Path(__file__).resolve().parent / "AISI_Manual_Test_Plan_P0.xlsx"

thin = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
header_fill = PatternFill("solid", fgColor="1B4F72")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
title_font = Font(bold=True, name="Calibri", size=16, color="1B4F72")
subtitle_font = Font(bold=True, name="Calibri", size=12, color="2874A6")
section_fill = PatternFill("solid", fgColor="D4E6F1")
p0_fill = PatternFill("solid", fgColor="F5B7B1")
p1_fill = PatternFill("solid", fgColor="AED6F1")
happy_fill = PatternFill("solid", fgColor="D5F5E3")
neg_fill = PatternFill("solid", fgColor="FADBD8")
smoke_fill = PatternFill("solid", fgColor="FCF3CF")
pass_fill = PatternFill("solid", fgColor="82E0AA")
fail_fill = PatternFill("solid", fgColor="F1948A")
blocked_fill = PatternFill("solid", fgColor="F7DC6F")
yes_fill = PatternFill("solid", fgColor="D5F5E3")
no_fill = PatternFill("solid", fgColor="F5B7B1")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin


def style_data_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin
        cell.alignment = wrap
        cell.font = Font(name="Calibri", size=10)


def add_case(cases, cid, modul, judul, tipe, prio, role, pre, steps, expected, data):
    cases.append(
        (cid, modul, judul, tipe, prio, role, pre, steps, expected, data, "Not Run", "", "", "")
    )


def build_p1_cases() -> list[tuple]:
    cases: list[tuple] = []

    # ========== EVT Agenda ==========
    add_case(
        cases, "EVT-01", "AGENDA", "SA/Admin buat event semua sekolah", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "Login SA atau Admin",
        "1. Buka Agenda (/events)\n2. Klik Tambah\n3. Isi Judul (≥2 karakter)\n"
        "4. Isi waktu Mulai & Berakhir (end > start)\n5. Isi Poin check-in\n"
        "6. Cakupan: Semua sekolah (atau pilih sekolah)\n"
        "7. Centang minimal satu level kelompok (LEVEL_1/LEVEL_2)\n"
        "8. Pastikan dipublikasikan\n9. Klik Buat event",
        "Event muncul di list Agenda; status/waktu sesuai; anggota eligible bisa melihat saat ongoing",
        "Judul unik + window waktu overlapping sekarang jika ingin uji check-in",
    )
    add_case(
        cases, "EVT-02", "AGENDA", "PJ buat event scope sekolah sendiri", "Happy", "P1",
        "PJ_SEKOLAH",
        "Login PJ SMAN 1",
        "1. Buka Agenda → Tambah\n2. Isi judul, waktu, poin, level\n"
        "3. Pastikan tidak perlu pilih sekolah lain (otomatis sekolah PJ)\n4. Buat event",
        "Event dibuat hanya untuk sekolah PJ; teks UI menjelaskan cakupan sekolah",
        "Event SMAN 1 saja",
    )
    add_case(
        cases, "EVT-03", "AGENDA", "Anggota check-in dengan foto (event ongoing)", "Happy", "P1",
        "ANGGOTA",
        "Ada event Sedang berlangsung untuk level/sekolah anggota",
        "1. Login Anggota\n2. Buka Agenda → buka event ongoing\n"
        "3. Ambil/pilih foto check-in\n4. Klik Kirim check-in\n5. Amati status",
        "Status Menunggu persetujuan pembina; foto terunggah; tidak dapat double-submit pending",
        "Foto JPG/PNG < ~5MB",
    )
    add_case(
        cases, "EVT-04", "AGENDA", "Pembina setujui check-in → poin masuk", "Happy", "P1",
        "PEMBINA",
        "Ada check-in PENDING dari anggota kelompoknya (EVT-03)",
        "1. Login Pembina\n2. Agenda → Persetujuan (/events/check-ins)\n"
        "3. Review foto & detail\n4. Klik Setujui\n5. Cek poin anggota bertambah sebesar pointValue event",
        "Status APPROVED; poin anggota naik; PointLog tercatat",
        "Anggota kelompok pembina",
    )
    add_case(
        cases, "EVT-05", "AGENDA", "Pembina tolak check-in → anggota bisa kirim ulang", "Happy", "P1",
        "PEMBINA → ANGGOTA",
        "Ada check-in PENDING",
        "1. Pembina: Tolak check-in\n2. Login Anggota\n3. Buka event yang sama\n4. Kirim check-in lagi dengan foto baru",
        "Setelah ditolak, form check-in tersedia lagi; submit ulang berhasil (pending baru)",
        "—",
    )
    add_case(
        cases, "EVT-N01", "AGENDA", "Anggota/Pembina ditolak buat event", "Negatif", "P1",
        "ANGGOTA, PEMBINA",
        "Login role terkait",
        "1. Pastikan tombol Tambah tidak ada di Agenda\n2. Akses langsung /events/new",
        "Redirect /dashboard atau akses ditolak",
        "—",
    )
    add_case(
        cases, "EVT-N02", "AGENDA", "Validasi level & judul & waktu event", "Negatif", "P1",
        "ADMIN",
        "/events/new",
        "1. Submit tanpa centang level → harus ditolak (\"Pilih minimal satu level\")\n"
        "2. Judul 1 karakter → ditolak\n3. End ≤ Start → ditolak",
        "Validasi FE/API menolak; event tidak dibuat",
        "Input invalid",
    )
    add_case(
        cases, "EVT-N03", "AGENDA", "Check-in tanpa foto / di luar window waktu", "Negatif", "P1",
        "ANGGOTA",
        "Event ongoing + event belum mulai / sudah berakhir",
        "1. Event ongoing: coba kirim tanpa foto\n"
        "2. Event belum mulai / sudah berakhir: pastikan form check-in tidak muncul",
        "Tanpa foto → error wajib foto; di luar window → tidak bisa check-in",
        "—",
    )
    add_case(
        cases, "EVT-N04", "AGENDA", "Double check-in APPROVED/PENDING ditolak", "Negatif", "P1",
        "ANGGOTA",
        "Sudah punya check-in PENDING atau APPROVED",
        "1. Coba kirim check-in kedua pada event yang sama",
        "Error: sudah check-in event ini",
        "—",
    )
    add_case(
        cases, "EVT-N05", "AGENDA", "Visibility event sekolah/level lain", "Negatif", "P1",
        "ANGGOTA",
        "Ada event untuk sekolah/level lain",
        "1. Login anggota SMAN 1 level tertentu\n2. Buka list Agenda\n3. Pastikan event sekolah/level tidak eligible tidak tampil",
        "Hanya event published + endAt≥now + school/level match yang terlihat",
        "—",
    )

    # ========== MAT Materi ==========
    add_case(
        cases, "MAT-01", "MATERI", "SA/Admin publikasikan materi (file/link/tulis)", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "Login SA/Admin",
        "1. Buka Materi (/materi) → Tambah\n"
        "2. Isi Judul, Ringkasan (opsional), Pekan (Senin), level target\n"
        "3. Pilih jenis: Upload file / Link / Tulis langsung\n"
        "4. Lengkapi konten sesuai jenis\n5. Klik Publikasikan materi",
        "Materi muncul di list; detail menampilkan konten sesuai tipe",
        "FILE: PDF kecil; LINK: URL valid; RICH_TEXT: paragraf singkat",
    )
    add_case(
        cases, "MAT-02", "MATERI", "Buka detail materi per tipe konten", "Happy", "P1",
        "PJ / PEMBINA / ADMIN",
        "Ada materi FILE, LINK, RICH_TEXT",
        "1. Buka list Materi\n2. Buka detail LINK → tombol Buka materi\n"
        "3. Detail FILE → unduh/lihat\n4. Detail RICH_TEXT → konten ter-render",
        "Tiap tipe menampilkan aksi/konten yang benar; tidak blank",
        "Materi dari MAT-01",
    )
    add_case(
        cases, "MAT-03", "MATERI", "PJ/Pembina baca list tanpa tombol Tambah", "Smoke", "P1",
        "PJ_SEKOLAH, PEMBINA",
        "Login role terkait",
        "1. Buka /materi\n2. Pastikan list load\n3. Pastikan tombol Tambah tidak ada\n4. Buka satu detail",
        "Bisa baca; tidak bisa create via UI",
        "—",
    )
    add_case(
        cases, "MAT-N01", "MATERI", "Non-admin ditolak buat materi", "Negatif", "P1",
        "PJ, PEMBINA, ANGGOTA",
        "Login role terkait",
        "1. Akses /materi/new via URL",
        "Redirect /dashboard",
        "—",
    )
    add_case(
        cases, "MAT-N02", "MATERI", "Validasi level & konten wajib", "Negatif", "P1",
        "ADMIN",
        "/materi/new",
        "1. Submit tanpa level\n2. Tipe LINK tanpa URL\n3. Tipe FILE tanpa file",
        "Validasi menolak; materi tidak terbit",
        "—",
    )
    add_case(
        cases, "MAT-N03", "MATERI", "Anggota tidak punya menu Materi", "Negatif", "P1",
        "ANGGOTA",
        "Login Anggota",
        "1. Cek sidebar — Materi tidak ada\n2. Opsional: akses /materi via URL (catat perilaku)",
        "Tidak di nav; jika URL dibuka, sesuai RoleGuard/API (catat hasil)",
        "—",
    )

    # ========== ANL Analitik ==========
    add_case(
        cases, "ANL-01", "ANALITIK", "SA/Admin lihat Analitik Depok", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "Login SA/Admin; ada data seed",
        "1. Buka Analitik (/analytics)\n"
        "2. Verifikasi kartu: Sekolah, Kelompok, Pembina, Anggota, submit rate\n"
        "3. Cek panel gender/level breakdown\n4. Cek tren kehadiran (jika ada)",
        "Dashboard analitik terisi; tidak blank/crash",
        "—",
    )
    add_case(
        cases, "ANL-02", "ANALITIK", "PJ lihat Analitik sekolah (scope)", "Happy", "P1",
        "PJ_SEKOLAH",
        "Login PJ",
        "1. Buka Analitik\n2. Pastikan judul/konteks sekolah (bukan kota global)\n"
        "3. Angka sesuai sekolah assign saja",
        "Data scoped ke sekolah PJ",
        "SMAN 1",
    )
    add_case(
        cases, "ANL-N01", "ANALITIK", "Pembina/Anggota ditolak Analitik", "Negatif", "P1",
        "PEMBINA, ANGGOTA",
        "Login role terkait",
        "1. Pastikan menu Analitik tidak ada\n2. Akses /analytics via URL",
        "Redirect /dashboard",
        "—",
    )

    # ========== MUT Mutabaah ==========
    add_case(
        cases, "MUT-01", "MUTABAAH", "Anggota simpan draft mutabaah pekan", "Happy", "P1",
        "ANGGOTA",
        "Login Anggota; master mutabaah sudah ada (seed/config)",
        "1. Buka Mutabaah (/mutabaah)\n2. Pilih Pekan (Senin)\n"
        "3. Jika multi-kelompok, pilih Kelompok\n4. Isi beberapa field\n5. Simpan draft",
        "Draft tersimpan; bisa dibuka lagi di pekan yang sama",
        "Pekan ≥ tanggal bergabung",
    )
    add_case(
        cases, "MUT-02", "MUTABAAH", "Anggota kirim mutabaah → +2 poin", "Happy", "P1",
        "ANGGOTA",
        "Draft siap / form terisi",
        "1. Lengkapi field wajib\n2. Klik Kirim mutabaah\n"
        "3. Amati banner sudah dikirim\n4. Cek poin bertambah +2",
        "Submitted; field disabled; poin +2",
        "Pekan yang belum di-submit",
    )
    add_case(
        cases, "MUT-03", "MUTABAAH", "Pembina lihat mutabaah anggota (read-only)", "Happy", "P1",
        "PEMBINA",
        "Anggota sudah isi mutabaah",
        "1. Login Pembina\n2. Buka detail kelompok → detail anggota\n"
        "3. Buka panel/section Mutabaah per pekan",
        "Data mutabaah tampil read-only; tidak bisa edit sebagai pembina di panel ini",
        "—",
    )
    add_case(
        cases, "MUT-N01", "MUTABAAH", "Non-anggota ditolak /mutabaah", "Negatif", "P1",
        "PEMBINA, ADMIN, PJ",
        "Login role terkait",
        "1. Pastikan Mutabaah tidak di sidebar\n2. Akses /mutabaah",
        "Redirect /dashboard",
        "—",
    )
    add_case(
        cases, "MUT-N02", "MUTABAAH", "Pekan sebelum bergabung / submit ulang", "Negatif", "P1",
        "ANGGOTA",
        "Sudah submit satu pekan",
        "1. Coba pilih pekan sebelum minWeekDate (bergabung) — harus diblok\n"
        "2. Buka pekan sudah submit — field disabled; submit ulang gagal",
        "Validasi pekan & submit sekali saja dihormati",
        "—",
    )

    # ========== IC ==========
    add_case(
        cases, "IC-01", "IC", "SA/Admin CRUD master Indikator Capaian", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "Login SA → Pengaturan → IC atau /config/ic",
        "1. Filter level/kategori jika ada\n"
        "2. Tambah IC: Kategori, Tipe (Primer/Sekunder), Nomor, Indikator, Materi\n"
        "3. Simpan\n4. Edit satu item\n5. Hapus item uji (bukan data kritis seed jika perlu)",
        "CRUD master sukses; list ter-update",
        "Nomor/judul unik uji",
    )
    add_case(
        cases, "IC-02", "IC", "PJ/Pembina baca master IC (view only)", "Smoke", "P1",
        "PJ_SEKOLAH, PEMBINA",
        "Login → menu Indikator Capaian",
        "1. Buka /config/ic\n2. Pastikan list IC tampil\n3. Pastikan form Tambah/Edit tidak tersedia (atau disabled)",
        "View only; tidak bisa mengubah master",
        "—",
    )
    add_case(
        cases, "IC-03", "IC", "Pembina centang progress IC anggota", "Happy", "P1",
        "PEMBINA",
        "Login pembina; buka detail anggota kelompoknya",
        "1. Expand kategori IC\n2. Toggle indikator tercapai\n"
        "3. Amati toast Progress IC diperbarui\n4. Cek ringkasan X/Y %",
        "Progress tersimpan; persentase update",
        "Anggota dengan level yang punya IC",
    )
    add_case(
        cases, "IC-N01", "IC", "PJ/non-pembina tidak boleh edit progress/master", "Negatif", "P1",
        "PJ_SEKOLAH, ANGGOTA",
        "Login role terkait",
        "1. PJ: pastikan tidak bisa Tambah IC di master\n"
        "2. Anggota: pastikan tidak melihat panel progress IC di UI-nya\n"
        "3. (Opsional API) POST progress sebagai non-pembina → 403",
        "403 / UI tidak menyediakan aksi",
        "—",
    )
    add_case(
        cases, "IC-N02", "IC", "Validasi nomor/indikator wajib saat tambah IC", "Negatif", "P1",
        "SUPERADMIN",
        "/config/ic form tambah",
        "1. Submit tanpa nomor atau tanpa indikator",
        "Pesan validasi; item tidak dibuat",
        "—",
    )

    # ========== KKS ==========
    add_case(
        cases, "KKS-01", "KKS", "User kirim Keluhan/Kritik/Saran", "Happy", "P1",
        "ANGGOTA / PEMBINA / PJ",
        "Login salah satu role non-admin",
        "1. Buka KKS (/kks)\n2. Pilih Jenis (Keluhan/Kritik/Saran)\n"
        "3. Isi Subjek (≥3 karakter) & Pesan (≥10 karakter)\n4. Kirim\n5. Pastikan muncul di riwayat sendiri",
        "KKS terkirim status Menunggu; muncul di list pengirim",
        "Subjek & pesan valid",
    )
    add_case(
        cases, "KKS-02", "KKS", "SA/Admin kelola status & catatan", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "Ada KKS PENDING dari KKS-01",
        "1. Buka KKS (kotak masuk)\n2. Filter Menunggu\n3. Buka detail\n"
        "4. Isi Catatan Admin\n5. Tandai Dibaca → Tandai Selesai\n6. Opsional: Kembalikan ke Menunggu",
        "Status berubah; catatan tersimpan; pengirim bisa lihat tanggapan",
        "—",
    )
    add_case(
        cases, "KKS-03", "KKS", "Pengirim melihat tanggapan admin", "Happy", "P1",
        "Pengirim KKS-01",
        "Admin sudah isi catatan / ubah status",
        "1. Login sebagai pengirim\n2. Buka detail KKS-nya\n3. Verifikasi Tanggapan Admin & status",
        "Tanggapan & status terlihat",
        "—",
    )
    add_case(
        cases, "KKS-N01", "KKS", "Validasi subjek/pesan terlalu pendek", "Negatif", "P1",
        "ANGGOTA",
        "/kks form",
        "1. Subjek 2 karakter / pesan 5 karakter → submit",
        "Validasi Zod/FE menolak",
        "—",
    )
    add_case(
        cases, "KKS-N02", "KKS", "Non-admin tidak akses KKS orang lain / ubah status", "Negatif", "P1",
        "ANGGOTA",
        "Ketahu ID KKS milik user lain",
        "1. Coba buka /kks/:id milik orang lain\n2. (API) PUT status → 403",
        "403 / tidak ditemukan",
        "—",
    )

    # ========== CFG Config ==========
    add_case(
        cases, "CFG-01", "CONFIG", "SA/Admin buka hub Pengaturan & simpan label level", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "Login SA/Admin → Pengaturan",
        "1. Buka /config\n2. Edit Label Level Kelompok jika ada\n3. Simpan\n4. Refresh — perubahan persist",
        "Pengaturan load; simpan sukses",
        "—",
    )
    add_case(
        cases, "CFG-02", "CONFIG", "Kelola master mutabaah", "Happy", "P1",
        "SUPERADMIN / ADMIN",
        "/config/mutabaah",
        "1. Pilih level\n2. Tambah poin: Judul, Tipe input, Cakupan waktu, Wajib\n"
        "3. Jika SELECT: isi minimal satu opsi\n4. Simpan\n5. Edit / hapus item uji",
        "Master mutabaah ter-update; anggota melihat field baru di pekan berikutnya",
        "Item uji non-kritis",
    )
    add_case(
        cases, "CFG-03", "CONFIG", "Link kelola IC dari Pengaturan", "Smoke", "P1",
        "SUPERADMIN",
        "/config",
        "1. Klik kelola indikator capaian\n2. Landasan /config/ic\n3. Konten tampil",
        "Navigasi config → IC lancar",
        "—",
    )
    add_case(
        cases, "CFG-N01", "CONFIG", "PJ/Pembina/Anggota ditolak /config & mutabaah master", "Negatif", "P1",
        "PJ, PEMBINA, ANGGOTA",
        "Login role terkait",
        "1. Akses /config dan /config/mutabaah via URL\n2. PJ/Pembina boleh /config/ic (view) — bedakan",
        "/config & mutabaah → redirect; /config/ic view OK untuk PJ/Pembina",
        "—",
    )
    add_case(
        cases, "CFG-N02", "CONFIG", "SELECT mutabaah tanpa opsi ditolak", "Negatif", "P1",
        "ADMIN",
        "Form tambah field SELECT",
        "1. Tipe SELECT tanpa pilihan\n2. Submit",
        "Error minimal satu pilihan wajib",
        "—",
    )

    # ========== NTF Notifikasi ==========
    add_case(
        cases, "NTF-01", "NOTIFIKASI", "List notifikasi unread/read", "Happy", "P1",
        "Semua",
        "Ada notifikasi (mis. dari KKS/undangan) atau seed",
        "1. Buka /notifications\n2. Verifikasi title, body, tanggal\n3. Bedakan unread vs read secara visual",
        "List load; empty state jelas jika kosong",
        "—",
    )
    add_case(
        cases, "NTF-02", "NOTIFIKASI", "Tandai semua dibaca", "Happy", "P1",
        "SUPERADMIN / role dengan unread",
        "unreadCount > 0",
        "1. Klik Tandai dibaca (semua)\n2. Refresh — unread hilang/berkurang",
        "Semua notifikasi marked read",
        "—",
    )
    add_case(
        cases, "NTF-N01", "NOTIFIKASI", "Empty state tidak ada notifikasi", "Smoke", "P1",
        "User tanpa notifikasi",
        "Akun baru / sudah dibaca semua",
        "1. Buka /notifications",
        "Pesan Tidak ada notifikasi (atau setara); tidak crash",
        "—",
    )

    # ========== LDR Leaderboard ==========
    add_case(
        cases, "LDR-01", "LEADERBOARD", "Direct URL leaderboard menampilkan ranking", "Smoke", "P1",
        "PEMBINA / ANGGOTA",
        "Ada poin dari evaluasi/event/mutabaah",
        "1. Buka langsung /leaderboard (tidak di sidebar)\n2. Verifikasi daftar ranking + poin",
        "Halaman load; ranking tampil (atau empty state)",
        "Catatan: route orphan — tidak di nav",
    )
    add_case(
        cases, "LDR-N01", "LEADERBOARD", "Tanpa auth → login", "Negatif", "P1",
        "Public",
        "Logout",
        "1. Buka /leaderboard",
        "Redirect /login",
        "—",
    )

    # ========== Cross-module journey P1 ==========
    add_case(
        cases, "E2E-P1-01", "E2E", "Journey: event check-in → approve → poin", "Happy", "P1",
        "ADMIN → ANGGOTA → PEMBINA",
        "Env siap; catat judul event uji",
        "1. Admin buat event ongoing (level & sekolah anggota)\n"
        "2. Anggota check-in + foto\n3. Pembina setujui\n4. Verifikasi poin anggota\n"
        "5. (Opsional) cek /leaderboard & notifikasi",
        "Rantai event→poin konsisten di UI",
        "Email/akun seed",
    )
    add_case(
        cases, "E2E-P1-02", "E2E", "Journey: config mutabaah → isi → submit → pembina lihat", "Happy", "P1",
        "ADMIN → ANGGOTA → PEMBINA",
        "Master mutabaah ada",
        "1. Admin pastikan master mutabaah untuk level anggota\n"
        "2. Anggota isi & kirim mutabaah pekan ini\n"
        "3. Pembina buka detail anggota → lihat mutabaah submitted + poin +2",
        "Data mutabaah & poin konsisten",
        "—",
    )

    return cases


def main() -> None:
    wb = load_workbook(OUT)
    # Idempotent: jangan double-append jika EVT-01 sudah ada
    ws_check = wb["04_Test_Cases"]
    existing = {
        ws_check.cell(row=r, column=1).value
        for r in range(5, ws_check.max_row + 1)
        if ws_check.cell(row=r, column=1).value
    }
    if "EVT-01" in existing:
        print("P1 cases already present (EVT-01 found). Skipping append.")
        return
    p1_cases = build_p1_cases()

    # ----- Update Petunjuk -----
    ws = wb["01_Petunjuk"]
    # Find last used row in col A
    last = ws.max_row
    r = last + 2
    ws.cell(row=r, column=1, value="Update fitur P1").font = Font(
        bold=True, name="Calibri", size=11, color="1B4F72"
    )
    ws.cell(row=r, column=1).fill = section_fill
    ws.cell(row=r, column=1).border = thin
    ws.cell(
        row=r,
        column=2,
        value=(
            f"Ditambah {date.today().isoformat()}: case P1 untuk AGENDA (EVT), MATERI (MAT), "
            "ANALITIK (ANL), MUTABAAH (MUT), IC, KKS, CONFIG (CFG), NOTIFIKASI (NTF), "
            "LEADERBOARD (LDR), plus journey E2E-P1. "
            "Filter kolom Prioritas = P1 di sheet 04_Test_Cases atau lihat sheet 08_Test_Cases_P1. "
            "Urutan saran: CFG → MUT/IC → EVT → MAT → ANL → KKS → NTF → LDR."
        ),
    ).alignment = wrap
    ws.cell(row=r, column=2).border = thin
    ws.row_dimensions[r].height = 70

    # Update cakupan row if present
    for row in range(1, last + 1):
        if ws.cell(row=row, column=1).value == "Cakupan P0":
            ws.cell(row=row, column=1).value = "Cakupan P0 + P1"
            old = str(ws.cell(row=row, column=2).value or "")
            if "AGENDA" not in old:
                ws.cell(row=row, column=2).value = (
                    old.replace(
                        "Modul Agenda, Mutabaah, IC, KKS, Analitik, Materi, Config detail → P1 "
                        "(kecuali smoke buka halaman config SUPERADMIN di nav loop).",
                        "P0: AUTH–EVAL inti. P1: Agenda, Materi, Analitik, Mutabaah, IC, KKS, "
                        "Config, Notifikasi, Leaderboard (lihat sheet P1 / filter Prioritas).",
                    )
                    if "P1" in old or "Agenda" in old
                    else old
                    + " | P1 ditambahkan: EVT/MAT/ANL/MUT/IC/KKS/CFG/NTF/LDR."
                )
            break

    # ----- Extend Matriks -----
    ws3 = wb["03_Matriks_Akses"]
    start = ws3.max_row + 2
    ws3.cell(row=start, column=1, value="Matriks akses fitur P1 (tambahan)").font = subtitle_font
    mh = [
        "Halaman / Fitur",
        "SUPERADMIN",
        "ADMIN",
        "PJ_SEKOLAH",
        "PEMBINA",
        "ANGGOTA",
        "Catatan Uji",
    ]
    for i, h in enumerate(mh, 1):
        ws3.cell(row=start + 1, column=i, value=h)
    style_header(ws3, start + 1, len(mh))
    matrix = [
        ("/events list", "Ya", "Ya", "Ya", "Ya", "Ya", "EVT"),
        ("Buat event", "Ya", "Ya", "Ya", "—", "—", "EVT-01/02"),
        ("Check-in foto", "—", "—", "—", "—", "Ya", "EVT-03"),
        ("Approve check-in", "—", "—", "—", "Ya", "—", "EVT-04"),
        ("/materi list", "Ya", "Ya", "Ya", "Ya", "—*", " *tidak di nav anggota"),
        ("Buat materi", "Ya", "Ya", "—", "—", "—", "MAT-01"),
        ("/analytics", "Ya", "Ya", "Scope", "—", "—", "ANL"),
        ("/mutabaah isi", "—", "—", "—", "—", "Ya", "MUT"),
        ("IC master edit", "Ya", "Ya", "lihat", "lihat", "—", "IC-01"),
        ("IC progress centang", "—", "—", "—", "Ya", "—", "IC-03"),
        ("/kks", "Ya", "Ya", "Ya", "Ya", "Ya", "KKS"),
        ("/config hub", "Ya", "Ya", "—", "—", "—", "CFG"),
        ("/config/mutabaah", "Ya", "Ya", "—", "—", "—", "CFG-02"),
        ("/notifications", "Ya", "Ya", "Ya", "Ya", "Ya", "NTF"),
        ("/leaderboard", "URL", "URL", "URL", "URL", "URL", "tidak di sidebar"),
    ]
    for i, m in enumerate(matrix):
        r = start + 2 + i
        for c, v in enumerate(m, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = center if 1 < c < 7 else wrap
            cell.font = Font(name="Calibri", size=10)
            if c in (2, 3, 4, 5, 6):
                if v == "Ya" or str(v).startswith("Ya"):
                    cell.fill = yes_fill
                elif v == "—":
                    cell.fill = no_fill

    # ----- Append to 04_Test_Cases -----
    ws4 = wb["04_Test_Cases"]
    # find last data row (Case ID in col A)
    last_data = 4
    for row in range(5, ws4.max_row + 1):
        if ws4.cell(row=row, column=1).value:
            last_data = row
    start_row = last_data + 1

    # blank separator note
    # actually just append cases
    for i, case in enumerate(p1_cases):
        r = start_row + i
        for c, v in enumerate(case, 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = wrap
            cell.font = Font(name="Calibri", size=9)
        tipe = case[3]
        if tipe == "Happy":
            ws4.cell(row=r, column=4).fill = happy_fill
        elif tipe == "Negatif":
            ws4.cell(row=r, column=4).fill = neg_fill
        else:
            ws4.cell(row=r, column=4).fill = smoke_fill
        prio = case[4]
        ws4.cell(row=r, column=5).fill = p1_fill if prio == "P1" else p0_fill
        ws4.cell(row=r, column=5).alignment = center
        ws4.cell(row=r, column=1).alignment = center
        ws4.cell(row=r, column=11).alignment = center
        steps_lines = case[7].count("\n") + 1
        ws4.row_dimensions[r].height = min(160, max(50, steps_lines * 12 + 20))

    end_row = start_row + len(p1_cases) - 1
    ws4.auto_filter.ref = f"A4:N{end_row}"

    # refresh status validation range
    dv = DataValidation(
        type="list",
        formula1='"Not Run,Pass,Fail,Blocked,Skip"',
        allow_blank=True,
    )
    ws4.add_data_validation(dv)
    dv.add(f"K{start_row}:K{end_row}")
    ws4.conditional_formatting.add(
        f"K{start_row}:K{end_row}", FormulaRule(formula=['$K5="Pass"'], fill=pass_fill)
    )
    # fix formula to use relative row - openpyxl uses first cell; OK for range
    ws4.conditional_formatting.add(
        f"K{start_row}:K{end_row}",
        FormulaRule(formula=[f'$K{start_row}="Pass"'], fill=pass_fill),
    )
    ws4.conditional_formatting.add(
        f"K{start_row}:K{end_row}",
        FormulaRule(formula=[f'$K{start_row}="Fail"'], fill=fail_fill),
    )
    ws4.conditional_formatting.add(
        f"K{start_row}:K{end_row}",
        FormulaRule(formula=[f'$K{start_row}="Blocked"'], fill=blocked_fill),
    )

    # ----- Dedicated P1 sheet (copy of new cases for easy focus) -----
    if "08_Test_Cases_P1" in wb.sheetnames:
        del wb["08_Test_Cases_P1"]
    ws8 = wb.create_sheet("08_Test_Cases_P1")
    ws8["A1"] = "Test Cases Manual P1 — Agenda, Materi, Analitik, Mutabaah, IC, KKS, Config, dll."
    ws8["A1"].font = title_font
    ws8.merge_cells("A1:N1")
    ws8["A2"] = (
        "Salinan case P1 agar mudah difilter. Status di sheet ini independen — "
        "utamakan update Status di 04_Test_Cases agar Ringkasan akurat, "
        "atau samakan keduanya."
    )
    ws8["A2"].font = Font(name="Calibri", size=9, italic=True)
    cols = [
        "Case ID", "Modul", "Judul Case", "Tipe", "Prioritas", "Role Penguji",
        "Prekondisi", "Langkah Uji (Cara Test)", "Hasil Diharapkan", "Data Uji",
        "Status", "Actual Result", "Tester", "Tanggal",
    ]
    for i, h in enumerate(cols, 1):
        ws8.cell(row=4, column=i, value=h)
    style_header(ws8, 4, len(cols))
    for i, case in enumerate(p1_cases):
        r = 5 + i
        for c, v in enumerate(case, 1):
            cell = ws8.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.alignment = wrap
            cell.font = Font(name="Calibri", size=9)
        tipe = case[3]
        ws8.cell(row=r, column=4).fill = (
            happy_fill if tipe == "Happy" else neg_fill if tipe == "Negatif" else smoke_fill
        )
        ws8.cell(row=r, column=5).fill = p1_fill
        ws8.cell(row=r, column=5).alignment = center
        ws8.row_dimensions[r].height = min(140, max(45, (case[7].count("\n") + 1) * 11 + 18))
    ws8.auto_filter.ref = f"A4:N{4 + len(p1_cases)}"
    ws8.freeze_panes = "A5"
    widths = [10, 12, 42, 10, 10, 22, 36, 55, 42, 28, 12, 28, 14, 12]
    for i, w in enumerate(widths, 1):
        ws8.column_dimensions[get_column_letter(i)].width = w
    dv8 = DataValidation(type="list", formula1='"Not Run,Pass,Fail,Blocked,Skip"', allow_blank=True)
    ws8.add_data_validation(dv8)
    dv8.add(f"K5:K{4 + len(p1_cases)}")

    # ----- Checklist P1 -----
    if "09_Checklist_P1" in wb.sheetnames:
        del wb["09_Checklist_P1"]
    ws9 = wb.create_sheet("09_Checklist_P1")
    ws9["A1"] = "Checklist Cepat P1 (45–60 menit)"
    ws9["A1"].font = title_font
    for i, h in enumerate(["#", "Checklist", "Role", "Done (✓)"], 1):
        ws9.cell(row=3, column=i, value=h)
    style_header(ws9, 3, 4)
    checklist = [
        ("Buat 1 event ongoing (Admin/PJ)", "Admin/PJ"),
        ("Anggota check-in + foto", "Anggota"),
        ("Pembina setujui → poin naik", "Pembina"),
        ("Publikasi 1 materi (link atau file)", "Admin"),
        ("PJ/Pembina buka detail materi", "PJ/Pembina"),
        ("Analitik SA + Analitik PJ load", "SA + PJ"),
        ("Anggota draft + submit mutabaah (+2 poin)", "Anggota"),
        ("Pembina lihat mutabaah di detail anggota", "Pembina"),
        ("SA tambah 1 IC master (atau edit)", "SA"),
        ("Pembina centang progress IC anggota", "Pembina"),
        ("User kirim KKS", "Anggota"),
        ("Admin tandai dibaca/selesai + catatan", "Admin"),
        ("Config mutabaah: buka + 1 perubahan kecil", "Admin"),
        ("Notifikasi: list + tandai dibaca", "SA"),
        ("/leaderboard load via URL langsung", "Anggota"),
    ]
    for i, (text, role) in enumerate(checklist, 1):
        r = 3 + i
        ws9.cell(row=r, column=1, value=i)
        ws9.cell(row=r, column=2, value=text)
        ws9.cell(row=r, column=3, value=role)
        ws9.cell(row=r, column=4, value="")
        style_data_row(ws9, r, 4)
        ws9.cell(row=r, column=1).alignment = center
    for i, w in enumerate([6, 55, 14, 12], 1):
        ws9.column_dimensions[get_column_letter(i)].width = w
    ws9.freeze_panes = "A4"

    # ----- Update Ringkasan formulas & modul list -----
    ws5 = wb["05_Ringkasan"]
    # Update Total formula to full range
    ws5["B4"] = f"=COUNTA('04_Test_Cases'!A5:A{end_row})"
    ws5["B5"] = f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Not Run\")"
    ws5["B6"] = f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Pass\")"
    ws5["B7"] = f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Fail\")"
    ws5["B8"] = f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Blocked\")"
    ws5["B9"] = f"=COUNTIF('04_Test_Cases'!K5:K{end_row},\"Skip\")"

    # Add P1 module rows after existing breakdown
    # Find "AUTH" row
    modul_start = None
    for row in range(1, ws5.max_row + 1):
        if ws5.cell(row=row, column=1).value == "AUTH":
            modul_start = row
            break
    if modul_start:
        # find last modul row
        r = modul_start
        while ws5.cell(row=r, column=1).value:
            r += 1
        p1_moduls = [
            ("AGENDA", sum(1 for c in p1_cases if c[1] == "AGENDA")),
            ("MATERI", sum(1 for c in p1_cases if c[1] == "MATERI")),
            ("ANALITIK", sum(1 for c in p1_cases if c[1] == "ANALITIK")),
            ("MUTABAAH", sum(1 for c in p1_cases if c[1] == "MUTABAAH")),
            ("IC", sum(1 for c in p1_cases if c[1] == "IC")),
            ("KKS", sum(1 for c in p1_cases if c[1] == "KKS")),
            ("CONFIG", sum(1 for c in p1_cases if c[1] == "CONFIG")),
            ("NOTIFIKASI", sum(1 for c in p1_cases if c[1] == "NOTIFIKASI")),
            ("LEADERBOARD", sum(1 for c in p1_cases if c[1] == "LEADERBOARD")),
        ]
        # Also count E2E P1 in E2E row or add
        existing_moduls = set()
        for row in range(modul_start, r):
            existing_moduls.add(ws5.cell(row=row, column=1).value)
        write_at = r
        for m, cnt in p1_moduls:
            if m in existing_moduls:
                continue
            ws5.cell(row=write_at, column=1, value=m)
            ws5.cell(row=write_at, column=2, value=cnt)
            ws5.cell(
                row=write_at,
                column=3,
                value=f"=COUNTIFS('04_Test_Cases'!B:B,A{write_at},'04_Test_Cases'!K:K,\"Pass\")",
            )
            ws5.cell(
                row=write_at,
                column=4,
                value=f"=COUNTIFS('04_Test_Cases'!B:B,A{write_at},'04_Test_Cases'!K:K,\"Fail\")",
            )
            ws5.cell(
                row=write_at,
                column=5,
                value=f"=COUNTIFS('04_Test_Cases'!B:B,A{write_at},'04_Test_Cases'!K:K,\"Not Run\")",
            )
            style_data_row(ws5, write_at, 6)
            write_at += 1

    # Note on P0 vs P1 counts
    note_row = ws5.max_row + 2
    ws5.cell(row=note_row, column=1, value="Catatan jumlah case").font = subtitle_font
    ws5.cell(
        row=note_row + 1,
        column=1,
        value=(
            f"P0 asli ~59 case. Ditambah {len(p1_cases)} case P1 pada {date.today().isoformat()}. "
            f"Total di 04_Test_Cases ≈ {last_data - 4 + len(p1_cases)}. "
            "Filter kolom Prioritas untuk fokus sesi P0 vs P1."
        ),
    ).alignment = wrap
    ws5.merge_cells(start_row=note_row + 1, start_column=1, end_row=note_row + 1, end_column=6)
    ws5.row_dimensions[note_row + 1].height = 45

    wb.save(OUT)
    print(f"Updated {OUT}")
    print(f"Added {len(p1_cases)} P1 cases")
    from collections import Counter
    print("P1 modules:", dict(Counter(c[1] for c in p1_cases)))


if __name__ == "__main__":
    main()
